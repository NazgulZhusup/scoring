import pandas as pd
import asyncpg
from fastapi.responses import FileResponse
from app.config import config
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


async def get_db_connection():
    return await asyncpg.connect(config.DATABASE_URL)


async def generate_excel(app_id: int):
    """
    Генерирует детализированный Excel-отчёт по заявке.
    """
    conn = await get_db_connection()
    query = "SELECT * FROM applications WHERE id = $1"
    application = await conn.fetchrow(query, app_id)
    await conn.close()

    if not application:
        return None  # Если заявки нет, возвращаем None

    # Создаём новую книгу Excel
    wb = Workbook()

    # Основные данные
    ws1 = wb.active
    ws1.title = "Основные данные"
    ws1.append(["Поле", "Значение"])
    for key, value in dict(application).items():
        ws1.append([key, value])

    # Лист расчёта скоринга
    ws2 = wb.create_sheet("Скоринговый анализ")
    ws2.append(["Фактор", "Баллы"])

    # Получаем детали скоринга
    details = eval(application["details"]) if application.get("details") else {}


    for factor, points in details.items():
        row = [factor, points]
        ws2.append(row)
        # Если риск высокий (баллы < 0), окрашиваем в красный
        if points < 0:
            for cell in ws2[ws2.max_row]:
                cell.fill = PatternFill(start_color="FF6666", end_color="FF6666", fill_type="solid")

    # Сохраняем в файл
    file_path = f"application_{app_id}.xlsx"
    wb.save(file_path)

    return file_path
